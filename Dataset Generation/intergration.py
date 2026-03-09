import openpyxl
import numpy as np
import os

def clear_integrated_images_folder():
    """
    清空 integrated_images 文件夹中的所有文件。
    """
    folder = 'integrated_images'
    if os.path.exists(folder):
        for file in os.listdir(folder):
            file_path = os.path.join(folder, file)
            if os.path.isfile(file_path):
                os.remove(file_path)
    else:
        os.makedirs(folder)

def get_num_files():
    """
    自动获取 images/excel 文件夹内的文件数量。
    """
    folder = 'images/excel'
    if os.path.exists(folder):
        return len([f for f in os.listdir(folder) if f.endswith('.xlsx')])
    return 0

def load_data(filename):
    # 使用 openpyxl 读取 .xlsx 文件
    book = openpyxl.load_workbook(filename)
    sheet1 = book['images']  # 获取名为 'images' 的工作表
    x = np.zeros((50, 50))  # 初始化50x50的矩阵
    for i in range(50):
        for j in range(50):
            # openpyxl 中的单元格索引是从1开始的
            cell_value = sheet1.cell(row=i + 1, column=j + 1).value
            if cell_value is not None:
                x[i, j] = cell_value
            else:
                x[i, j] = 0

    return x.reshape(1, 2500)  # 返回一个形状为(1, 2500)的数组

def generate_soil_parameters(n):
    Es = np.round(np.random.uniform(1, 100, n), 2).reshape(-1, 1)
    Ps = np.round(np.random.uniform(0.3, 0.45, n), 3).reshape(-1, 1)
    rhos = np.round(np.random.uniform(1.5, 2.2, n), 3).reshape(-1, 1)
    params = np.concatenate([Es, Ps, rhos], axis=1)
    return params


def save_data(x, params, num, filename):
    # 使用 openpyxl 写入数据到新的 excel 文件
    book = openpyxl.Workbook()  # 创建一个新的工作簿
    sheet1 = book.active
    sheet1.title = 'images'  # 创建名为 'images' 的工作表
    sheet2 = book.create_sheet('soil_parameters')  # 创建名为 'soil_parameters' 的工作表

    print(f"Saving file: {filename}")  # 确保文件名正确
    print(f"x shape: {x.shape}, params shape: {params.shape}")  # 打印数组形状

    for i in range(num):
        for j in range(50 * 50):
            sheet1.cell(row=i + 1, column=j + 1, value=float(x[i, j]))  # 确保数据被转换成数值存储
        for j in range(params.shape[1]):
            sheet2.cell(row=i + 1, column=j + 1, value=float(params[i, j]))  # 确保数据正确存储

    book.save(filename)
    print(f"File saved: {filename}")  # 确保文件被正确保存

clear_integrated_images_folder()

total_files = get_num_files()
num = total_files
#num = 10
total_batches = total_files // num

for m in range(total_batches):
    x = np.zeros((num, 2500))
    for n in range(num):
        t = load_data('images/excel/'+str(m*num+n)+'.xlsx')
        x[n] = t[0]
    params = generate_soil_parameters(num)
    save_data(x, params, num, 'integrated_images/'+str(m)+'.xlsx')
